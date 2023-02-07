from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.title='HCL'
'''sheet['A1'].value=1
sheet['B1'].value='python'
sheet.cell(row=1,column=3).value='write'
wb.save('../data/demo_workbook_write.xlsx')'''
j=0
for i in range(10,60,10):
    j+=1
    #for j in range(1,6):
    sheet.cell(row=1,column=j).value=i
wb.save('../data/demo_workbook_write.xlsx')