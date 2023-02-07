import openpyxl
wb=openpyxl.load_workbook('../data/Book1.xlsx')
sheet=wb.active
data=sheet['A3'].value

data1=sheet.cell(row=2,column=2).value
data2=sheet['A1:A10']
print(data)
print(data1)
print(data2)
print(sheet.max_row)
print(sheet.max_cloumn)